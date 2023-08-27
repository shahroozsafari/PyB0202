
from os import system
clear = lambda : system("cls")


from pickle import load
try:
    with open("books.info","rb") as books_info:
        books = load(books_info)
except(FileNotFoundError):
    books= []

def add_book():
    clear()
    global books                # global list of all books (optional code)
    book = {} 
    book["title"] = input("Enter title of the book : ")
    book["author"] = input("Enter author of the book : ")
    try:
        book["pages"] = int(input("Enter pages of the book : "))
        book["price"] = float(input("Enter price of the book : "))
    except(ValueError):
        input("\n\nPages and Price MUST be a number, Press any key to retrun to Menu ...")
        return False
    book["isbn"] = input("Enter ISBN of the book : ")
    if check_isbn(book["isbn"]) :
        input("This book already exists in books database !")
        return False
    books.append(book)
    input("Press any key ...")

def list_books():
    clear()
    for book in books :
        print(f"Title : {book['title']}")
        print(f"Author : {book['author']}")
        print(f"Pages : {book['pages']}")
        print(f"Price : {book['price']}")
        print(f"ISBN : {book['isbn']}")
        print("----------------------------------")
    input("Press any key ...")
    # print(books.__sizeof__())

def find_book():
    clear()
    isbn = input("Enter ISBN to find your book :")
    for book in books :
        if book["isbn"] == isbn :
            print("----------------------------------")
            print(f"Title : {book['title']}")
            print(f"Author : {book['author']}")
            print(f"Pages : {book['pages']}")
            print(f"Price : {book['price']}")
            print(f"ISBN : {book['isbn']}")
            print("----------------------------------")
            input("Press any key ...")
            break
    else :
        input("\n\nThis book IS NOT in books database !")


def delete_book():
    clear()
    isbn = input("Enter ISBN to delete book :")
    for book in books :
        if book["isbn"] == isbn :
            books.remove(book)
            input("Book has been deleted successfully ")
            break
    else :
        input("This book IS NOT in books database !")

def save_books():
    clear()
    from pickle import dump
    try:
        with open("books.info","wb") as books_info:
            dump(books,books_info)
            input("Books has been saved successfully ")
    except(PermissionError):
        input("\n\nPlease select another location to save books database")

def check_isbn(isbn):
    for book in books:
        if book["isbn"] == isbn or len(isbn) != 5 :
            return True
    return False












#------------------- Functions to send to Excel---------
def send_to_excel(index=0,**book_data):
    import openpyxl as xl
    wb = xl.load_workbook("Books_Excel.xlsx")
    ws = wb.active
    index=ws.cell(1,6).value
    ws.cell(index,1).value = book_data["title"]
    ws.cell(index,2).value = book_data["author"]
    ws.cell(index,3).value = book_data["pages"]
    ws.cell(index,4).value = book_data["price"]
    ws.cell(index,5).value = book_data["isbn"]
    index+=1
    ws.cell(1,6).value = index
    wb.save("Books_Excel.xlsx")

def get_from_excel():
    import openpyxl as xl
    wb = xl.load_workbook("Books_Excel.xlsx")
    ws = wb.active
    index=ws.cell(1,6).value
    for i in range(2,index):
        book={}
        book["title"] = ws.cell(i,1).value
        book["author"] = ws.cell(i,2).value
        book["pages"] = ws.cell(i,3).value
        book["price"] = ws.cell(i,4).value
        book["isbn"] = ws.cell(i,5).value
        books.append(book)